"""Загрузка ниш из SQLite вместо xlsx."""
import sqlite3
import os
from dataclasses import dataclass
from config import DATA_FILE

DB_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "wb_trends.db")


@dataclass
class Niche:
    query: str
    requests: int
    products: int
    competition: float


def _get_db_path() -> str:
    path = os.path.normpath(DB_PATH)
    if not os.path.exists(path):
        # Fallback на старый xlsx
        return None
    return path


def load_niches(filepath: str = None) -> list[Niche]:
    """Загружает свободные ниши из SQLite (конкуренция ≤5, запросы ≥500)."""
    db_path = _get_db_path()
    if db_path is None:
        # Fallback на старый xlsx
        return _load_niches_xlsx(filepath)
    
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute(
        "SELECT phrase, request_count, cards_count, competition "
        "FROM niches WHERE competition <= 5 AND request_count >= 500 "
        "ORDER BY request_count DESC"
    )
    niches = []
    for row in c:
        niches.append(Niche(
            query=row["phrase"],
            requests=row["request_count"],
            products=row["cards_count"],
            competition=row["competition"],
        ))
    conn.close()
    return niches


def get_categories(niches: list[Niche]) -> list[str]:
    """Возвращает уникальные категории из БД."""
    db_path = _get_db_path()
    if db_path is None:
        return list(set(n.category for n in niches if hasattr(n, 'category') and n.category))
    
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("SELECT DISTINCT category FROM niches WHERE category IS NOT NULL ORDER BY category")
    cats = [row[0] for row in c.fetchall()]
    conn.close()
    return cats


def filter_by_category(niches: list[Niche], category: str) -> list[Niche]:
    """Фильтрует ниши по категории через SQLite."""
    db_path = _get_db_path()
    if db_path is None:
        return [n for n in niches if hasattr(n, 'category') and n.category == category]
    
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute(
        "SELECT phrase, request_count, cards_count, competition "
        "FROM niches WHERE category = ? AND competition <= 5 AND request_count >= 500 "
        "ORDER BY request_count DESC",
        (category,)
    )
    result = []
    for row in c:
        result.append(Niche(
            query=row["phrase"],
            requests=row["request_count"],
            products=row["cards_count"],
            competition=row["competition"],
        ))
    conn.close()
    return result


def filter_by_keywords(niches: list[Niche], keywords: list[str]) -> list[Niche]:
    """Фильтрует ниши по ключевым словам в запросе."""
    kw_lower = [k.lower() for k in keywords]
    return [n for n in niches if any(k in n.query.lower() for k in kw_lower)]


def format_niche(n: Niche) -> str:
    """Форматирует нишу для вывода."""
    comp = f"{n.competition:.1f}" if n.competition < 100 else "∞"
    return f"📋 {n.query} · 📊 {n.requests:,} запросов · 📦 {n.products:,} товаров · 🎯 конкуренция {comp}"


def wb_search_url(query: str) -> str:
    """Генерирует ссылку на поиск WB."""
    from urllib.parse import quote
    return f"https://www.wildberries.ru/catalog/0/search.aspx?query={quote(query)}"


def _load_niches_xlsx(filepath: str) -> list[Niche]:
    """Старый fallback — загрузка из xlsx."""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    niches = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 4:
            query = str(row[0] or "")
            requests = int(row[1] or 0)
            products = int(row[2] or 0)
            competition = float(row[3] or 0)
            if query:
                niches.append(Niche(query, requests, products, competition))
    wb.close()
    return niches